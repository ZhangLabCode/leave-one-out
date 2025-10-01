# version12_lightgrey_highcontrast.py
# Light–gray HIGH-CONTRAST UI (white cards, blue accent). Clean checkboxes (no label boxes).
# Workflow: Import/Stack → Filter → Fit & LOO → Predict/Plots, with Prev/Next + clickable step chips.

from __future__ import annotations
from pathlib import Path
from typing import Optional, Tuple, List

import numpy as np
import pandas as pd

from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtCore import Qt, Signal

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.widgets import LassoSelector, RectangleSelector
from matplotlib.path import Path as MplPath
from matplotlib import rcParams
from cycler import cycler

APP_TITLE = "UV-Vis • Stack → Fit → LOO → Predict"
DEFAULT_X = "A_ISARS"
DEFAULT_Y = "UV-Vis"
DEFAULT_GROUP = "Wavelength"
DEFAULT_CONC = "Concentration"

# ---------------- THEME ----------------
def apply_light_hicontrast_theme(app: QtWidgets.QApplication):
    QtWidgets.QApplication.setStyle("Fusion")

    # Modern font
    for fam in ["Inter", "Segoe UI Variable", "Segoe UI", "Arial"]:
        f = QtGui.QFont(fam, 10)
        if QtGui.QFontInfo(f).family():
            app.setFont(f); break

    # Palette (light gray + blue accent, HIGH contrast)
    blue      = QtGui.QColor("#2563EB")   # accent
    blue_alt  = QtGui.QColor("#1D4ED8")
    bg        = QtGui.QColor("#F5F7FA")   # page background (light gray)
    card      = QtGui.QColor("#FFFFFF")   # cards / panes
    field     = QtGui.QColor("#FFFFFF")   # inputs (white)
    line      = QtGui.QColor("#CBD5E1")   # outlines
    text      = QtGui.QColor("#0F172A")   # near-black text
    midtext   = QtGui.QColor("#111827")
    subtext   = QtGui.QColor("#475569")
    altrow    = QtGui.QColor("#F1F5F9")

    pal = QtGui.QPalette()
    CR = QtGui.QPalette.ColorRole
    CG = QtGui.QPalette.ColorGroup
    for cg in (CG.Active, CG.Inactive, CG.Disabled):
        pal.setColor(cg, CR.Window,          bg)
        pal.setColor(cg, CR.WindowText,      text)
        pal.setColor(cg, CR.Base,            field)
        pal.setColor(cg, CR.AlternateBase,   altrow)
        pal.setColor(cg, CR.ToolTipBase,     card)
        pal.setColor(cg, CR.ToolTipText,     text)
        pal.setColor(cg, CR.Text,            text)
        pal.setColor(cg, CR.Button,          card)
        pal.setColor(cg, CR.ButtonText,      text)
        pal.setColor(cg, CR.Highlight,       blue)
        pal.setColor(cg, CR.HighlightedText, Qt.white)
    app.setPalette(pal)

    # Global stylesheet
    #  * All labels/checkboxes have transparent backgrounds (kills unwanted rectangles)
    #  * Checkboxes get a clean, modern indicator
    #  * Tabs are text-only with a blue underline on selection
    app.setStyleSheet(f"""
        QWidget {{
            background: {bg.name()};
            color: {text.name()};
            font-size: 10.5pt;
        }}

        /* Transparent labels = no gray boxes anywhere */
        QLabel {{ background: transparent; }}
        QLabel[hint="true"] {{ color: {subtext.name()}; font-size: 9.7pt; }}

        /* Tooltips */
        QToolTip {{
            background: {card.name()};
            color: {text.name()};
            border: 1px solid {line.name()};
            padding: 6px 8px;
        }}

        /* Header card */
        #Header {{
            background: {card.name()};
            border: 1px solid {line.name()};
            border-radius: 12px;
            padding: 10px;
        }}

        /* Cards/Groups */
        QGroupBox {{
            background: {card.name()};
            border: 1px solid {line.name()};
            border-radius: 12px;
            margin-top: 14px;
            padding: 10px 10px 6px 10px;
        }}
        QGroupBox::title {{
            subcontrol-origin: margin;
            left: 12px;
            padding: 0 6px;
            color: {midtext.name()};
            font-weight: 800;
            background: transparent;
        }}

        /* Inputs */
        QLineEdit, QComboBox {{
            background: {field.name()};
            border: 1px solid {line.name()};
            border-radius: 10px;
            padding: 7px 10px;
            color: {text.name()};
        }}
        QLineEdit:focus, QComboBox:focus {{ border: 2px solid {blue.name()}; }}
        QComboBox::drop-down {{ border: none; width: 24px; }}
        QComboBox::down-arrow {{ width: 10px; height: 10px; }}

        /* Checkboxes – transparent label + modern indicator */
        QCheckBox {{
            background: transparent;  /* removes label box */
            spacing: 8px;
            color: {text.name()};
        }}
        /* Indicator with visible checkmark */
        QCheckBox::indicator {{
            width: 18px; height: 18px;
            border: 2px solid {line.name()};
            border-radius: 4px;
            background: {card.name()};
        }}
        QCheckBox::indicator:hover {{
            border-color: {blue.name()};
        }}
        QCheckBox::indicator:checked {{
            background: {blue.name()};
            border-color: {blue.name()};
            image: url(:/qt-project.org/styles/commonstyle/images/checkboxindicator.png);
        }}
        QCheckBox::indicator:unchecked {{
            image: none;
        }}

        /* Buttons */
        QPushButton {{
            background: {card.name()};
            border: 1px solid {line.name()};
            border-radius: 10px;
            padding: 8px 14px;
            font-weight: 700;
            color: {text.name()};
        }}
        QPushButton:hover {{ background: {altrow.name()}; }}
        QPushButton[accent="true"] {{
            background: {blue.name()};
            color: white;
            border: none;
        }}
        QPushButton[accent="true"]:hover {{ background: {blue_alt.name()}; }}

        /* Tabs – text only + blue underline when selected */
        QTabWidget::pane {{
            border: 1px solid {line.name()};
            border-radius: 12px;
            top: -1px;
            background: {card.name()};
        }}
        QTabBar::tab {{
            background: transparent;
            border: none;
            padding: 10px 18px;
            margin: 0 6px;
            color: {midtext.name()};
            font-weight: 800;
        }}
        QTabBar::tab:hover     {{ color: {text.name()}; }}
        QTabBar::tab:selected  {{
            color: {text.name()};
            border-bottom: 3px solid {blue.name()};
            margin-bottom: -3px;
        }}

        /* Tables */
        QTableWidget {{
            background: {card.name()};
            gridline-color: {line.name()};
            alternate-background-color: {altrow.name()};
            selection-background-color: {blue.name()};
            selection-color: white;
            border: 1px solid {line.name()};
            border-radius: 10px;
        }}
        QHeaderView::section {{
            background: {bg.name()};
            border: 1px solid {line.name()};
            padding: 7px;
            font-weight: 800;
            color: {midtext.name()};
        }}

        /* Scrollbars (clean, light) */
        QScrollBar:vertical {{
            background: transparent; width: 12px; margin: 2px; border-radius: 6px;
        }}
        QScrollBar::handle:vertical {{
            background: {line.name()}; min-height: 24px; border-radius: 6px;
        }}
        QScrollBar::handle:vertical:hover {{ background: {blue_alt.name()}; }}
        QScrollBar:horizontal {{
            background: transparent; height: 12px; margin: 2px; border-radius: 6px;
        }}
        QScrollBar::handle:horizontal {{
            background: {line.name()}; min-width: 24px; border-radius: 6px;
        }}
        QScrollBar::handle:horizontal:hover {{ background: {blue_alt.name()}; }}

        QStatusBar {{
            background: {card.name()};
            border: 1px solid {line.name()};
            border-radius: 10px;
            color: {subtext.name()};
        }}

        /* Step chips transparent */
        QFrame[role="chip"], QFrame[role="chip"] * {{ background: transparent; }}
    """)

    # Matplotlib – light, high contrast
    rcParams.update({
        "figure.facecolor": card.name(),
        "axes.facecolor": card.name(),
        "savefig.facecolor": card.name(),
        "axes.edgecolor": "#1F2937",
        "axes.labelcolor": "#0F172A",
        "xtick.color": "#475569",
        "ytick.color": "#475569",
        "grid.color": "#E2E8F0",
        "grid.linestyle": "-",
        "grid.alpha": 0.6,
        "axes.prop_cycle": cycler(color=[
            "#2563EB", "#0EA5E9", "#16A34A", "#7C3AED", "#EA580C", "#DB2777"
        ]),
        "font.size": 10.5,
    })

# -------------- HELPERS --------------
def df_to_table(tbl: QtWidgets.QTableWidget, df: pd.DataFrame, max_rows: int = 400):
    tbl.clear()
    if df is None or df.empty:
        tbl.setRowCount(0); tbl.setColumnCount(0); return
    view = df.head(max_rows)
    tbl.setRowCount(len(view)); tbl.setColumnCount(len(view.columns))
    tbl.setHorizontalHeaderLabels([str(c) for c in view.columns])
    for r in range(len(view)):
        for c, col in enumerate(view.columns):
            v = view.iloc[r, c]
            item = QtWidgets.QTableWidgetItem("" if pd.isna(v) else f"{v}")
            item.setTextAlignment(Qt.AlignCenter)
            tbl.setItem(r, c, item)
    tbl.horizontalHeader().setStretchLastSection(True)
    tbl.resizeColumnsToContents()

def centered_r2(y: np.ndarray, yhat: np.ndarray) -> float:
    sse = np.sum((y - yhat) ** 2)
    sst = np.sum((y - y.mean()) ** 2)
    return float(1.0 - (sse / sst if sst > 0 else np.nan))

def safe_float(s: str):
    s = (s or "").strip()
    try: return float(s) if s else None
    except ValueError: return None


def _parse_wavelength_exclusions(text: str):
    """Parse a user string like "200-220, 350, 500-505" into a list of (lo, hi) tuples.
    Singles are treated as (v, v). Ignores blanks and malformed tokens.
    """
    import re as _re  # local import to ensure availability
    out = []
    if not text:
        return out
    for tok in _re.split(r"[,\s]+", text.strip()):
        if not tok:
            continue
        if "-" in tok:
            parts = tok.split("-")
            try:
                lo = float(parts[0]); hi = float(parts[-1])
                if lo > hi:
                    lo, hi = hi, lo
                out.append((lo, hi))
            except ValueError:
                continue
        else:
            try:
                v = float(tok)
                out.append((v, v))
            except ValueError:
                continue
    return out

class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, w=6.8, h=4.0, dpi=120):
        self.fig = Figure(figsize=(w, h), dpi=dpi)
        self.ax = self.fig.add_subplot(111)
        super().__init__(self.fig)
        self.setParent(parent)

# ---------- Step Chip ----------
class StepChip(QtWidgets.QFrame):
    clicked = Signal(int)
    def __init__(self, number: int, title: str):
        super().__init__()
        self.setProperty("role", "chip")
        self.setAutoFillBackground(False)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self.index = number - 1

        self.circle = QtWidgets.QLabel(str(number))
        self.circle.setFixedSize(28, 28)
        self.circle.setAlignment(Qt.AlignCenter)

        self.text = QtWidgets.QLabel(title)
        self.text.setStyleSheet("font-weight:800;")

        lay = QtWidgets.QHBoxLayout(self)
        lay.setContentsMargins(4, 0, 4, 0)
        lay.setSpacing(8)
        lay.addWidget(self.circle); lay.addWidget(self.text)

        self.setCursor(Qt.PointingHandCursor)
        self._apply(False)

    def _apply(self, active: bool):
        if active:
            self.circle.setStyleSheet("border-radius:14px; background:#2563EB; color:white; font-weight:900;")
            self.text.setStyleSheet("color:#0F172A; font-weight:900;")
        else:
            self.circle.setStyleSheet("border-radius:14px; background:#E5E7EB; color:#111827; font-weight:900;")
            self.text.setStyleSheet("color:#334155; font-weight:800;")

    def set_active(self, active: bool): self._apply(active)
    def mousePressEvent(self, e: QtGui.QMouseEvent):
        if e.button() == Qt.LeftButton: self.clicked.emit(self.index)
        super().mousePressEvent(e)

# -------------- MAIN WINDOW --------------
class MainWindow(QtWidgets.QMainWindow):

    def _predicted_uv_by_concentration(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Return a pivot table:
            rows    = Wavelength
            columns = Concentration
            values  = UV-Vis_pred
        """
        if df is None or df.empty:
            return pd.DataFrame()

        t = df.copy()
        # Ensure numeric sort for axis labels
        for col in ["Wavelength", "Concentration", "UV-Vis_pred"]:
            if col in t.columns:
                t[col] = pd.to_numeric(t[col], errors="coerce")

        piv = (t.pivot_table(index="Wavelength",
                            columns="Concentration",
                            values="UV-Vis_pred",
                            aggfunc="first")
                .sort_index()            # wavelength ascending
                .sort_index(axis=1))     # concentration ascending
        # Make Wavelength a normal column again
        piv = piv.reset_index()
        return piv

    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1260, 840)

        # state
        self.path_A: Optional[str] = None
        self.path_U: Optional[str] = None
        self.stacked_df: Optional[pd.DataFrame] = None
        self.filtered_df: Optional[pd.DataFrame] = None
        self.loo_df: Optional[pd.DataFrame] = None
        self.pred_df: Optional[pd.DataFrame] = None
        self.fit_coeffs: Optional[Tuple[float,float,float]] = None
        self.fit_r2: Optional[float] = None
        self.fit_meta = {}

        # layout root
        central = QtWidgets.QWidget()
        root = QtWidgets.QVBoxLayout(central)
        root.setContentsMargins(14, 14, 14, 12)
        root.setSpacing(12)

        # Header
        head = QtWidgets.QFrame(objectName="Header")
        head_l = QtWidgets.QHBoxLayout(head)
        head_l.setContentsMargins(12, 10, 12, 10)
        head_l.setSpacing(12)

        title_box = QtWidgets.QVBoxLayout()
        title = QtWidgets.QLabel("UV-Vis Modeling"); title.setStyleSheet("font-size: 20pt; font-weight: 900;")
        subtitle = QtWidgets.QLabel("Import & stack → Clean → Fit & LOO → Predict & visualize")
        subtitle.setProperty("hint","true")
        title_box.addWidget(title); title_box.addWidget(subtitle)

        self.step1 = StepChip(1, "Data")
        self.step2 = StepChip(2, "Fit & LOO")
        self.step3 = StepChip(3, "Predict & Plots")
        for ch in (self.step1, self.step2, self.step3): ch.clicked.connect(self._go_to_step)

        chips = QtWidgets.QHBoxLayout()
        chips.setSpacing(16)
        chips.addWidget(self.step1); chips.addWidget(self.step2); chips.addWidget(self.step3)
        chips.addStretch(1)

        head_l.addLayout(title_box, 1)
        head_l.addLayout(chips, 1)
        root.addWidget(head)

        # Tabs
        self.tabs = QtWidgets.QTabWidget()
        root.addWidget(self.tabs, 1)

        self._build_tab_data()
        self._build_tab_fit()
        self._build_tab_predict()

        # Footer nav
        footer = QtWidgets.QHBoxLayout()
        footer.setSpacing(10)
        self.btn_prev = QtWidgets.QPushButton("← Previous")
        self.btn_next = QtWidgets.QPushButton("Next →"); self.btn_next.setProperty("accent", True)
        footer.addStretch(1); footer.addWidget(self.btn_prev); footer.addWidget(self.btn_next)
        root.addLayout(footer)

        self.setCentralWidget(central)

        # wire up
        self.tabs.currentChanged.connect(self._on_tab_changed)
        self.btn_prev.clicked.connect(self._prev_tab)
        self.btn_next.clicked.connect(self._next_tab)

        self._on_tab_changed(0)
        self.statusBar().showMessage("Step 1: Load A_ISARS & UV-Vis, then click Combine.")

    # Navigation
    def _go_to_step(self, index: int): self.tabs.setCurrentIndex(index)
    def _prev_tab(self): self.tabs.setCurrentIndex(max(0, self.tabs.currentIndex()-1))
    def _next_tab(self):
        i = self.tabs.currentIndex()
        self.tabs.setCurrentIndex(i+1 if i < self.tabs.count()-1 else i)
        if i == self.tabs.count()-1:
            QtWidgets.QMessageBox.information(self, "Done", "Workflow complete. Export or revisit any step.")

    def _on_tab_changed(self, idx: int):
        self.step1.set_active(idx == 0)
        self.step2.set_active(idx == 1)
        self.step3.set_active(idx == 2)
        self.btn_prev.setEnabled(idx > 0)
        self.btn_next.setText("Finish" if idx == self.tabs.count()-1 else "Next →")

    # ---------- Tab 1: Data ----------
    def _build_tab_data(self):
        tab = QtWidgets.QWidget(); v = QtWidgets.QVBoxLayout(tab)

        gb_imp = QtWidgets.QGroupBox("Step 1 — Import & Stack")
        iv = QtWidgets.QVBoxLayout(gb_imp)

        row = QtWidgets.QHBoxLayout()
        self.btn_open_A = QtWidgets.QPushButton("Open A_ISARS…"); self.btn_open_A.setProperty("accent", True)
        self.btn_open_U = QtWidgets.QPushButton("Open UV-Vis…");  self.btn_open_U.setProperty("accent", True)
        self.lbl_A = QtWidgets.QLabel("No A_ISARS file")
        self.lbl_U = QtWidgets.QLabel("No UV-Vis file")
        self.btn_combine = QtWidgets.QPushButton("Combine & Preview"); self.btn_combine.setProperty("accent", True)
        self.btn_export_stacked = QtWidgets.QPushButton("Export Stacked…")
        for w in (self.btn_open_A, self.lbl_A, self.btn_open_U, self.lbl_U, self.btn_combine, self.btn_export_stacked):
            row.addWidget(w)
        row.addStretch(1); iv.addLayout(row)

        lbl_hint = QtWidgets.QLabel("Matrices: wavelengths in Column A, concentrations in Row 1. Then click Combine.")
        lbl_hint.setProperty("hint","true"); iv.addWidget(lbl_hint)

        gb_f = QtWidgets.QGroupBox("Optional — Clean & Filter")
        fv = QtWidgets.QVBoxLayout(gb_f)
        self.chk_drop_nan = QtWidgets.QCheckBox("Drop rows where A_ISARS or UV-Vis is NaN"); self.chk_drop_nan.setChecked(True)
        self.chk_ranges = QtWidgets.QCheckBox("Enable numeric ranges for A_ISARS / UV-Vis")
        fv.addWidget(self.chk_drop_nan); fv.addWidget(self.chk_ranges)

        grid = QtWidgets.QGridLayout()
        grid.addWidget(QtWidgets.QLabel("A_ISARS min:"), 0, 0)
        self.ed_a_min = QtWidgets.QLineEdit(); self.ed_a_min.setEnabled(False); grid.addWidget(self.ed_a_min, 0, 1)
        grid.addWidget(QtWidgets.QLabel("A_ISARS max:"), 0, 2)
        self.ed_a_max = QtWidgets.QLineEdit(); self.ed_a_max.setEnabled(False); grid.addWidget(self.ed_a_max, 0, 3)
        grid.addWidget(QtWidgets.QLabel("UV-Vis min:"), 1, 0)
        self.ed_u_min = QtWidgets.QLineEdit(); self.ed_u_min.setEnabled(False); grid.addWidget(self.ed_u_min, 1, 1)
        grid.addWidget(QtWidgets.QLabel("UV-Vis max:"), 1, 2)
        self.ed_u_max = QtWidgets.QLineEdit(); self.ed_u_max.setEnabled(False); grid.addWidget(self.ed_u_max, 1, 3)
        fv.addLayout(grid)

        # Wavelength exclusions (comma/range list)
        wl_row = QtWidgets.QHBoxLayout()
        wl_row.addWidget(QtWidgets.QLabel("Exclude wavelengths (comma or ranges):"))
        self.ed_wl_exclude = QtWidgets.QLineEdit()
        self.ed_wl_exclude.setPlaceholderText("e.g., 200-220, 350, 500-505")
        wl_row.addWidget(self.ed_wl_exclude, 1)
        fv.addLayout(wl_row)

        lbl_wl_hint = QtWidgets.QLabel("Exclude exact values or ranges before fitting/LOO (applies to stacked data).")
        lbl_wl_hint.setProperty("hint","true")
        fv.addWidget(lbl_wl_hint)

        hb = QtWidgets.QHBoxLayout()
        self.btn_apply = QtWidgets.QPushButton("Apply filters"); self.btn_reset = QtWidgets.QPushButton("Reset")
        hb.addWidget(self.btn_apply); hb.addWidget(self.btn_reset); hb.addStretch(1)
        fv.addLayout(hb)

        v.addWidget(gb_imp); v.addWidget(gb_f)
        self.tbl_data = QtWidgets.QTableWidget(); self.tbl_data.setAlternatingRowColors(True)
        v.addWidget(self.tbl_data, 1)
        self.tabs.addTab(tab, "1) Data")

        # signals
        self.btn_open_A.clicked.connect(lambda: self._pick_file("A"))
        self.btn_open_U.clicked.connect(lambda: self._pick_file("U"))
        self.btn_combine.clicked.connect(self.combine_stack)
        self.btn_export_stacked.clicked.connect(self.export_stacked)
        self.chk_ranges.toggled.connect(self._toggle_ranges)
        self.btn_apply.clicked.connect(self.apply_filters)
        self.btn_reset.clicked.connect(self.reset_filters)

    def _toggle_ranges(self, checked: bool):
        for w in (self.ed_a_min, self.ed_a_max, self.ed_u_min, self.ed_u_max):
            w.setEnabled(checked)

    # ---------- Tab 2: Fit & LOO ----------
    def _build_tab_fit(self):
        tab = QtWidgets.QWidget(); v = QtWidgets.QVBoxLayout(tab)

        row = QtWidgets.QHBoxLayout()
        row.addWidget(QtWidgets.QLabel("X:"))
        self.cb_x = QtWidgets.QComboBox(); self.cb_x.setMinimumWidth(160); row.addWidget(self.cb_x)
        row.addWidget(QtWidgets.QLabel("Y:"))
        self.cb_y = QtWidgets.QComboBox(); self.cb_y.setMinimumWidth(160); row.addWidget(self.cb_y)
        row.addWidget(QtWidgets.QLabel("Group (LOO):"))
        self.cb_group = QtWidgets.QComboBox(); self.cb_group.setMinimumWidth(140); row.addWidget(self.cb_group)
        row.addWidget(QtWidgets.QLabel("Concentration:"))
        self.cb_conc = QtWidgets.QComboBox(); self.cb_conc.setMinimumWidth(140); row.addWidget(self.cb_conc)
        row.addStretch(1); v.addLayout(row)

        lbl_fit = QtWidgets.QLabel("‘Fit (ALL)’ gives one global quadratic. ‘Run LOO’ builds per-wavelength coefficients.")
        lbl_fit.setProperty("hint","true"); v.addWidget(lbl_fit)

        row2 = QtWidgets.QHBoxLayout()
        self.btn_fit = QtWidgets.QPushButton("Fit (ALL)"); self.btn_fit.setProperty("accent", True)
        self.btn_save_fig = QtWidgets.QPushButton("Save plot…")
        self.btn_export_coeffs = QtWidgets.QPushButton("Export coeffs…")
        row2.addWidget(self.btn_fit); row2.addWidget(self.btn_save_fig); row2.addWidget(self.btn_export_coeffs)
        row2.addSpacing(22)
        self.btn_loo = QtWidgets.QPushButton("Run LOO"); self.btn_loo.setProperty("accent", True)
        self.btn_export_loo = QtWidgets.QPushButton("Export LOO…")
        row2.addWidget(self.btn_loo); row2.addWidget(self.btn_export_loo)
        row2.addStretch(1); v.addLayout(row2)

        # Outlier selection controls
        row3 = QtWidgets.QHBoxLayout()
        self.btn_pick = QtWidgets.QPushButton("Pick points")
        self.btn_box = QtWidgets.QPushButton("Box select")
        self.btn_lasso = QtWidgets.QPushButton("Lasso select")
        self.btn_clear_sel = QtWidgets.QPushButton("Clear selection")
        self.btn_delete_sel = QtWidgets.QPushButton("Delete selected"); self.btn_delete_sel.setProperty("danger", True)
        self.lbl_sel = QtWidgets.QLabel("Selected: 0")
        row3.addWidget(self.btn_pick); row3.addWidget(self.btn_box); row3.addWidget(self.btn_lasso)
        row3.addSpacing(12); row3.addWidget(self.btn_clear_sel); row3.addWidget(self.btn_delete_sel)
        row3.addStretch(1); row3.addWidget(self.lbl_sel)
        v.addLayout(row3)

        self.lbl_eq = QtWidgets.QLabel(""); self.lbl_eq.setStyleSheet("font-weight:800;")
        v.addWidget(self.lbl_eq)

        self.canvas = MplCanvas(self, w=7.6, h=4.3, dpi=120)
        v.addWidget(self.canvas, 1)
        # Selection state for Fit/LOO
        self._sel_ids = set()
        self._sel_overlay = None
        self._last_fit_points = None
        self._pick_cid = None
        self._rect_selector = None
        self._lasso_selector = None
        self.tabs.addTab(tab, "2) Fit & LOO")

        # signals
        self.btn_fit.clicked.connect(self.run_fit)
        self.btn_save_fig.clicked.connect(self.save_fig)
        self.btn_export_coeffs.clicked.connect(self.export_coeffs)
        self.btn_loo.clicked.connect(self.run_loo)
        self.btn_export_loo.clicked.connect(self.export_loo)
        self.btn_pick.clicked.connect(self.enable_pick_mode)
        self.btn_box.clicked.connect(self.enable_box_mode)
        self.btn_lasso.clicked.connect(self.enable_lasso_mode)
        self.btn_clear_sel.clicked.connect(self.clear_selection)
        self.btn_delete_sel.clicked.connect(self.delete_selected)

    # ---------- Tab 3: Predict & Plots ----------
    def _build_tab_predict(self):
        tab = QtWidgets.QWidget(); v = QtWidgets.QVBoxLayout(tab)

        row = QtWidgets.QHBoxLayout()
        row.addWidget(QtWidgets.QLabel("Predict UV-Vis for"))
        self.cb_predict_group = QtWidgets.QComboBox(); self.cb_predict_group.setMinimumWidth(120); row.addWidget(self.cb_predict_group)
        self.btn_pred_sel = QtWidgets.QPushButton("Predict selected"); self.btn_pred_sel.setProperty("accent", True)
        self.btn_pred_all = QtWidgets.QPushButton("Predict ALL"); self.btn_pred_all.setProperty("accent", True)
        self.btn_export_pred = QtWidgets.QPushButton("Export preds…")
        self.btn_plots = QtWidgets.QPushButton("Spectra plots…")
        row.addWidget(self.btn_pred_sel); row.addWidget(self.btn_pred_all); row.addWidget(self.btn_export_pred); row.addWidget(self.btn_plots)
        row.addStretch(1); v.addLayout(row)

        lbl_pred = QtWidgets.QLabel("LOO coefficients per wavelength × A_ISARS → predicted UV-Vis.")
        lbl_pred.setProperty("hint","true"); v.addWidget(lbl_pred)

        self.tbl_pred = QtWidgets.QTableWidget(); self.tbl_pred.setAlternatingRowColors(True)
        v.addWidget(self.tbl_pred, 1)

        self.tabs.addTab(tab, "3) Predict & Plots")

        # signals
        self.btn_pred_sel.clicked.connect(self.predict_selected)
        self.btn_pred_all.clicked.connect(self.predict_all)
        self.btn_export_pred.clicked.connect(self.export_predictions)
        self.btn_plots.clicked.connect(self.show_spectra_plots)

    # ---------- Data I/O ----------
    def _pick_file(self, kind: str):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, f"Open {'A_ISARS' if kind=='A' else 'UV-Vis'} matrix", "",
            "Excel (*.xlsx *.xls);;CSV (*.csv)")
        if not path: return
        if kind == "A":
            self.path_A = path; self.lbl_A.setText(Path(path).name)
        else:
            self.path_U = path; self.lbl_U.setText(Path(path).name)
        self.statusBar().showMessage("Files selected. Click Combine.")

    @staticmethod
    def _read_matrix(path: str) -> pd.DataFrame:
        ext = Path(path).suffix.lower()
        if ext in (".xlsx", ".xls"): return pd.read_excel(path, sheet_name=0, header=None)
        return pd.read_csv(path, header=None)

    @staticmethod
    def _stack(df: pd.DataFrame, valname: str) -> pd.DataFrame:
        waves = pd.to_numeric(df.iloc[1:, 0], errors="coerce")

        header = pd.to_numeric(df.iloc[0, 1:], errors="coerce")
        header_has_numbers = header.notna().sum() >= max(1, len(header)//2)

        if header_has_numbers:
            reps = header.to_numpy()
            data = df.iloc[1:, 1:]
        else:
            # No numeric concentrations in the first row → use column indices 1..N
            data_all = df.iloc[1:, 1:]
            nonempty_cols = [
                i for i in range(data_all.shape[1])
                if pd.to_numeric(data_all.iloc[:, i], errors="coerce").notna().any()
            ]
            data = data_all.iloc[:, nonempty_cols]
            reps = np.arange(1, data.shape[1] + 1, dtype=float)

        vals = pd.to_numeric(data.to_numpy().reshape(-1), errors="coerce")
        W = np.repeat(waves.to_numpy(), repeats=len(reps))
        C = np.tile(reps, reps=len(waves))

        out = pd.DataFrame({"Wavelength": W, "Concentration": C, valname: vals})
        return out.dropna(subset=["Wavelength", "Concentration"])


    def combine_stack(self):
        if not self.path_A or not self.path_U:
            QtWidgets.QMessageBox.information(self, "Need files", "Select both A_ISARS and UV-Vis matrices."); return
        try:
            A = self._read_matrix(self.path_A); U = self._read_matrix(self.path_U)
            long_a = self._stack(A, "A_ISARS"); long_u = self._stack(U, "UV-Vis")
            merged = pd.merge(long_a, long_u, on=["Wavelength","Concentration"], how="outer", validate="one_to_one")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Combine error", str(e)); return

        merged = merged.sort_values(["Wavelength","Concentration"]).reset_index(drop=True)
        self.stacked_df = merged[["Wavelength","Concentration","A_ISARS","UV-Vis"]].copy()
        # Stable row IDs for manual exclusions / selections
        self.stacked_df["_id"] = np.arange(len(self.stacked_df), dtype=int)
        self.filtered_df = self.stacked_df.copy()
        self.manual_exclusions = set()
        df_to_table(self.tbl_data, self.filtered_df)

        cols = list(self.filtered_df.columns)
        for cb in (self.cb_x, self.cb_y, self.cb_group, self.cb_conc, self.cb_predict_group):
            cb.clear(); cb.addItems([str(c) for c in cols])

        def set_if(cb, name, fallback=0):
            i = cb.findText(name); cb.setCurrentIndex(i if i>=0 else min(fallback, cb.count()-1))
        set_if(self.cb_x, DEFAULT_X, 2)
        set_if(self.cb_y, DEFAULT_Y, 3)
        set_if(self.cb_group, DEFAULT_GROUP, 0)
        set_if(self.cb_conc, DEFAULT_CONC, 1)
        self._refresh_predict_groups()
        # Clear wavelength exclusions UI
        if hasattr(self, 'ed_wl_exclude'):
            self.ed_wl_exclude.clear()
        # Clear any manual point exclusions
        self.manual_exclusions = set()
        self.statusBar().showMessage(f"Combined • Rows: {len(self.filtered_df):,}")

    def export_stacked(self):
        if self.filtered_df is None or self.filtered_df.empty:
            QtWidgets.QMessageBox.information(self, "Nothing to export", "Combine (and filter) first."); return
        self._save_table(self.filtered_df, "Stacked")

    def apply_filters(self):
        if self.stacked_df is None:
            QtWidgets.QMessageBox.information(self, "No data", "Combine first."); return
        df = self.stacked_df.copy(); start = len(df)
        dropped_wl = 0
        if self.chk_drop_nan.isChecked():
            df = df.loc[~(df["A_ISARS"].isna() | df["UV-Vis"].isna())]
        if self.chk_ranges.isChecked():
            a_min = safe_float(self.ed_a_min.text()); a_max = safe_float(self.ed_a_max.text())
            u_min = safe_float(self.ed_u_min.text()); u_max = safe_float(self.ed_u_max.text())
            if a_min is not None: df = df.loc[~(df["A_ISARS"].notna() & (df["A_ISARS"] < a_min))]
            if a_max is not None: df = df.loc[~(df["A_ISARS"].notna() & (df["A_ISARS"] > a_max))]
            if u_min is not None: df = df.loc[~(df["UV-Vis"].notna() & (df["UV-Vis"] < u_min))]
            if u_max is not None: df = df.loc[~(df["UV-Vis"].notna() & (df["UV-Vis"] > u_max))]

        # Exclude wavelengths specified by the user (exact values or ranges)
        wl_text = self.ed_wl_exclude.text().strip() if hasattr(self, "ed_wl_exclude") else ""
        if wl_text:
            ranges = _parse_wavelength_exclusions(wl_text)
            if ranges:
                before = len(df)
                keep_mask = np.ones(len(df), dtype=bool)
                W = pd.to_numeric(df["Wavelength"], errors="coerce").to_numpy()
                for lo, hi in ranges:
                    in_range = (W >= lo) & (W <= hi)
                    if lo == hi:
                        in_range = (np.isfinite(W)) & (np.abs(W - lo) <= 1e-9)
                    keep_mask &= ~in_range
                df = df.loc[keep_mask]
                dropped_wl = before - len(df)
        # Drop any manually excluded rows (by stable _id)
        if hasattr(self, "manual_exclusions") and self.manual_exclusions:
            before2 = len(df)
            df = df.loc[~df["_id"].isin(list(self.manual_exclusions))]
            # dropped_manual = before2 - len(df)
        self.filtered_df = df.reset_index(drop=True)
        df_to_table(self.tbl_data, self.filtered_df)
        self._refresh_predict_groups()
        self.statusBar().showMessage(
            f"Filters applied • kept {len(self.filtered_df):,} rows (dropped {start-len(self.filtered_df):,})." +
            (f" Excluded {dropped_wl:,} by wavelength." if dropped_wl else "")
        )

    def reset_filters(self):
        if self.stacked_df is None: return
        self.filtered_df = self.stacked_df.copy()
        df_to_table(self.tbl_data, self.filtered_df)
        self._refresh_predict_groups()
        self.statusBar().showMessage(f"Filters reset • {len(self.filtered_df):,} rows.")

    # ---------- Fit & LOO ----------
    def _clean_xy(self, df, xcol, ycol):
        d = df.copy()
        d[xcol] = pd.to_numeric(d[xcol], errors="coerce")
        d[ycol] = pd.to_numeric(d[ycol], errors="coerce")
        d = d.replace([np.inf, -np.inf], np.nan).dropna(subset=[xcol, ycol])
        return d

    def run_fit(self):
        if self.filtered_df is None or self.filtered_df.empty:
            QtWidgets.QMessageBox.information(self, "No data", "Combine/filter first."); return
        xcol, ycol = self.cb_x.currentText(), self.cb_y.currentText()
        if not xcol or not ycol or xcol == ycol:
            QtWidgets.QMessageBox.information(self, "Pick columns", "Choose distinct X and Y."); return
        d = self._clean_xy(self.filtered_df[[xcol,ycol,"_id"]], xcol, ycol)
        if len(d) < 3:
            QtWidgets.QMessageBox.information(self, "Not enough data", "Need ≥3 rows."); return

        x = d[xcol].to_numpy(float); y = d[ycol].to_numpy(float)
        a,b,c = np.polyfit(x, y, 2); p = np.poly1d([a,b,c])
        yhat = p(x); r2 = centered_r2(y, yhat)
        self.fit_coeffs = (a,b,c); self.fit_r2 = r2; self.fit_meta = {"xcol":xcol,"ycol":ycol,"rows_used":len(d)}

        ax = self.canvas.ax
        ax.clear(); ax.grid(True, alpha=0.6)
        ax.set_title("Scatter and quadratic fit (ALL filtered rows)")
        ax.set_xlabel(xcol); ax.set_ylabel(ycol)
        sc = ax.scatter(x, y, s=16, alpha=0.7, picker=5)
        self._last_fit_points = {"x": x, "y": y, "ids": d["_id"].to_numpy(int), "scatter": sc}
        self._update_selection_overlay()
        xs = np.linspace(x.min(), x.max(), 500); ax.plot(xs, p(xs), linewidth=2.4)
        self.canvas.draw()
        self.lbl_eq.setText(f"y = {a:+.6f}x² {b:+.6f}x {c:+.6f}    R²={r2:.6f}    rows={len(d)}")
        self.statusBar().showMessage("Fit complete.")

    
    # ---------- Outlier selection mechanics (Fit tab) ----------
    def _disconnect_selectors(self):
        if self._rect_selector is not None:
            try: self._rect_selector.disconnect_events()
            except Exception: pass
            self._rect_selector = None
        if self._lasso_selector is not None:
            try: self._lasso_selector.disconnect_events()
            except Exception: pass
            self._lasso_selector = None
        if self._pick_cid is not None:
            try: self.canvas.mpl_disconnect(self._pick_cid)
            except Exception: pass
            self._pick_cid = None

    def _update_sel_label(self):
        if hasattr(self, "lbl_sel"):
            self.lbl_sel.setText(f"Selected: {len(self._sel_ids)}")

    def _update_selection_overlay(self):
        # Draw a hollow overlay on selected points
        if self._last_fit_points is None: return
        ax = self.canvas.ax
        if self._sel_overlay is not None:
            try: self._sel_overlay.remove()
            except Exception: pass
            self._sel_overlay = None
        if not self._sel_ids:
            self.canvas.draw(); self._update_sel_label(); return
        ids = self._last_fit_points["ids"]
        x = self._last_fit_points["x"]; y = self._last_fit_points["y"]
        mask = np.isin(ids, np.fromiter(self._sel_ids, dtype=int))
        self._sel_overlay = ax.scatter(x[mask], y[mask], s=80, facecolors="none", edgecolors="red", linewidths=1.8)
        self.canvas.draw()
        self._update_sel_label()

    def enable_pick_mode(self):
        # Single-point click selection (toggle)
        self._disconnect_selectors()
        if self._last_fit_points is None: return
        def on_pick(event):
            if event.artist != self._last_fit_points["scatter"]: return
            ind = int(event.ind[0]) if len(event.ind) else None
            if ind is None: return
            pid = int(self._last_fit_points["ids"][ind])
            if pid in self._sel_ids: self._sel_ids.remove(pid)
            else: self._sel_ids.add(pid)
            self._update_selection_overlay()
        self._pick_cid = self.canvas.mpl_connect("pick_event", on_pick)
        self.statusBar().showMessage("Pick mode: click points to toggle selection.")

    def enable_box_mode(self):
        # Drag a rectangle to add points to selection
        self._disconnect_selectors()
        if self._last_fit_points is None: return
        ax = self.canvas.ax
        def onselect(eclick, erelease):
            x0, y0 = eclick.xdata, eclick.ydata; x1, y1 = erelease.xdata, erelease.ydata
            if None in (x0,y0,x1,y1): return
            xmin, xmax = (x0, x1) if x0<=x1 else (x1, x0)
            ymin, ymax = (y0, y1) if y0<=y1 else (y1, y0)
            x = self._last_fit_points["x"]; y = self._last_fit_points["y"]; ids = self._last_fit_points["ids"]
            mask = (x>=xmin)&(x<=xmax)&(y>=ymin)&(y<=ymax)
            for pid in ids[mask]: self._sel_ids.add(int(pid))
            self._update_selection_overlay()
        self._rect_selector = RectangleSelector(ax, onselect, useblit=True, button=[1], interactive=False)
        self.statusBar().showMessage("Box select: drag to select outliers.")

    def enable_lasso_mode(self):
        # Free-form lasso to add points to selection
        self._disconnect_selectors()
        if self._last_fit_points is None: return
        ax = self.canvas.ax
        x = self._last_fit_points["x"]; y = self._last_fit_points["y"]; ids = self._last_fit_points["ids"]
        pts = np.column_stack([x, y])
        def onselect(verts):
            path = MplPath(verts)
            mask = path.contains_points(pts)
            for pid in ids[mask]: self._sel_ids.add(int(pid))
            self._update_selection_overlay()
        self._lasso_selector = LassoSelector(ax, onselect)
        self.statusBar().showMessage("Lasso select: draw around points to select.")

    def clear_selection(self):
        self._sel_ids = set(); self._update_selection_overlay()
        self.statusBar().showMessage("Selection cleared.")

    def delete_selected(self):
        if not self._sel_ids:
            QtWidgets.QMessageBox.information(self, "Nothing selected", "Use Pick/Box/Lasso to select points first."); return
        # Add to manual exclusions and re-apply filters
        if not hasattr(self, "manual_exclusions"):
            self.manual_exclusions = set()
        self.manual_exclusions.update(self._sel_ids)
        self._sel_ids = set(); self._last_fit_points = None
        self.apply_filters()
        # Optionally, auto-rerun the fit with same X/Y if still set
        try:
            self.run_fit()
        except Exception:
            pass
        self.statusBar().showMessage("Deleted selected outliers and refreshed.")
    def run_loo(self):
        if self.filtered_df is None or self.filtered_df.empty:
            QtWidgets.QMessageBox.information(self, "No data", "Combine/filter first."); return
        xcol, ycol, gcol = self.cb_x.currentText(), self.cb_y.currentText(), self.cb_group.currentText()
        if not xcol or not ycol or not gcol or xcol == ycol:
            QtWidgets.QMessageBox.information(self, "Pick columns", "Choose X, Y, and Group."); return

        base = self._clean_xy(self.filtered_df[[xcol,ycol,gcol]], xcol, ycol)
        if len(base) < 3:
            QtWidgets.QMessageBox.information(self, "Not enough data", "Need ≥3 rows after cleaning."); return

        groups = [g for g in pd.unique(self.filtered_df[gcol]) if pd.notna(g)]
        if not groups:
            QtWidgets.QMessageBox.information(self, "No groups", f"No non-NaN values in {gcol}."); return

        results = []
        for i, gv in enumerate(groups, 1):
            mask_remove = (base[gcol].astype(str) == str(gv))
            d_fit = base.loc[~mask_remove, [xcol, ycol]]
            rows_removed = int(mask_remove.sum()); rows_used = int(len(d_fit))
            if rows_used < 3:
                rec = {"group_removed": gv, "degree": 2, "a": np.nan, "b": np.nan, "c": np.nan, "R2": np.nan,
                       "rows_used": rows_used, "rows_removed": rows_removed}
            else:
                x = d_fit[xcol].to_numpy(float); y = d_fit[ycol].to_numpy(float)
                a,b,c = np.polyfit(x, y, 2); p = np.poly1d([a,b,c])
                r2 = centered_r2(y, p(x))
                rec = {"group_removed": gv, "degree": 2, "a": a, "b": b, "c": c, "R2": r2,
                       "rows_used": rows_used, "rows_removed": rows_removed}
            results.append(rec)
            if i % 25 == 0 or i == len(groups):
                self.statusBar().showMessage(f"LOO progress: {i}/{len(groups)}…")
                QtWidgets.QApplication.processEvents()

        self.loo_df = pd.DataFrame(results)
        self._refresh_predict_groups()
        self.statusBar().showMessage(f"LOO complete: {len(self.oo_df) if hasattr(self,'oo_df') else len(self.loo_df)} rows.")  # safe status text

    def save_fig(self):
        if self.fit_coeffs is None:
            QtWidgets.QMessageBox.information(self, "Nothing to save", "Run a fit first."); return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Save plot", "", "PNG (*.png);;SVG (*.svg);;PDF (*.pdf)")
        if not path: return
        self.canvas.fig.savefig(path, bbox_inches="tight", dpi=220)
        QtWidgets.QMessageBox.information(self, "Saved", f"Figure saved:\n{path}")

    def export_coeffs(self):
        if self.fit_coeffs is None or not self.fit_meta:
            QtWidgets.QMessageBox.information(self, "No results", "Run a fit first."); return
        a,b,c = self.fit_coeffs; r2 = self.fit_r2; meta = self.fit_meta
        row = pd.DataFrame([{"x_column":meta["xcol"], "y_column":meta["ycol"], "degree":2,
                             "a":a, "b":b, "c":c, "R2":r2, "rows_used":meta["rows_used"]}])
        self._save_table(row, "Coefficients")

    def export_loo(self):
        if self.loo_df is None or self.loo_df.empty:
            QtWidgets.QMessageBox.information(self, "No LOO", "Run LOO first."); return
        cols = ["group_removed","degree","a","b","c","R2","rows_used","rows_removed"]
        self._save_table(self.loo_df[cols], "LOO_Matrix")

    # ---------- Predict & Plots ----------
    def _refresh_predict_groups(self):
        self.cb_predict_group.clear()
        if self.filtered_df is None or self.filtered_df.empty: return
        gcol = self.cb_group.currentText()
        if not gcol or gcol not in self.filtered_df.columns: return
        vals = pd.unique(self.filtered_df[gcol].dropna())
        self.cb_predict_group.addItems([str(v) for v in vals])

    def _pick_group_rows(self, gval: str):
        xcol,ycol,gcol,ccol = self.cb_x.currentText(), self.cb_y.currentText(), self.cb_group.currentText(), self.cb_conc.currentText()
        keep = [c for c in [xcol,ycol,gcol,ccol] if c and c in self.filtered_df.columns]
        d = self.filtered_df[keep].copy()
        d[xcol] = pd.to_numeric(d[xcol], errors="coerce")
        d[ycol] = pd.to_numeric(d[ycol], errors="coerce")
        d = d.loc[d[gcol].astype(str) == str(gval)].dropna(subset=[xcol])
        return d,xcol,ycol,gcol,(ccol if ccol in d.columns else None)

    def _predict_block(self, d_sel, xcol, ycol, gcol, gval, ccol):
        if self.loo_df is None or self.loo_df.empty:
            return None
        row = self.loo_df.loc[self.loo_df["group_removed"].astype(str) == str(gval)]
        if row.empty or row[["a","b","c"]].isna().any(axis=None):
            return None

        a,b,c = float(row["a"].iloc[0]), float(row["b"].iloc[0]), float(row["c"].iloc[0])
        x = d_sel[xcol].to_numpy(float)
        y_pred = a*x**2 + b*x + c

        keep = [col for col in [gcol, xcol, ycol, ccol] if col]
        block = d_sel[keep].copy()
        block["UV-Vis_pred"] = y_pred

        # Compute RSD% = 100 * SD(pair) / mean(pair), where the pair is [actual, predicted]
        # SD for two values = |a - p| / sqrt(2); mean = (a + p) / 2
        actual = pd.to_numeric(block[ycol], errors="coerce")
        pred   = pd.to_numeric(block["UV-Vis_pred"], errors="coerce")

        mean_pair = (actual + pred) / 2.0
        sd_pair   = np.abs(actual - pred) / np.sqrt(2.0)

        # Avoid divide-by-zero; use NaN where mean is 0 or missing
        denom = np.where(np.abs(mean_pair) > 0, np.abs(mean_pair), np.nan)
        block["RSD%"] = (sd_pair / denom) * 100.0


        # Keep coefficients used for traceability
        block["a_used"], block["b_used"], block["c_used"] = a, b, c

        # Final column names
        ren = {gcol: "Wavelength", xcol: "A_ISARS", ycol: "UV-Vis_actual"}
        if ccol:
            ren[ccol] = "Concentration"
        return block.rename(columns=ren)


    def predict_selected(self):
        if self.filtered_df is None or self.filtered_df.empty or self.loo_df is None or self.loo_df.empty:
            QtWidgets.QMessageBox.information(self, "Missing", "Combine/filter and run LOO first."); return
        gval = self.cb_predict_group.currentText()
        d_sel,xcol,ycol,gcol,ccol = self._pick_group_rows(gval)
        if d_sel.empty: QtWidgets.QMessageBox.information(self,"No rows", f"No rows where {gcol}={gval}."); return
        block = self._predict_block(d_sel, xcol, ycol, gcol, gval, ccol)
        if block is None: QtWidgets.QMessageBox.information(self,"No coeffs", f"No valid LOO coefficients for {gval}."); return
        self.pred_df = block.reset_index(drop=True)
        df_to_table(self.tbl_pred, self.pred_df)
        self.statusBar().showMessage(f"Predicted {len(self.pred_df)} rows at {gcol}={gval}")

    def predict_all(self):
        if self.filtered_df is None or self.filtered_df.empty or self.loo_df is None or self.loo_df.empty:
            QtWidgets.QMessageBox.information(self, "Missing", "Combine/filter and run LOO first."); return
        xcol,ycol,gcol,ccol = self.cb_x.currentText(), self.cb_y.currentText(), self.cb_group.currentText(), self.cb_conc.currentText()
        groups = pd.unique(self.filtered_df[gcol].dropna())
        out: List[pd.DataFrame] = []
        for gv in groups:
            d_sel,_,_,_,_ = self._pick_group_rows(str(gv))
            if d_sel.empty: continue
            block = self._predict_block(d_sel, xcol, ycol, gcol, gv, (ccol if ccol in d_sel.columns else None))
            if block is not None: out.append(block)
        if not out:
            QtWidgets.QMessageBox.information(self, "No predictions", "Could not form predictions."); return
        self.pred_df = pd.concat(out, ignore_index=True)
        df_to_table(self.tbl_pred, self.pred_df)
        self.statusBar().showMessage(f"Predicted ALL groups • rows={len(self.pred_df):,}")

    def export_predictions(self):
        if self.pred_df is None or self.pred_df.empty:
            QtWidgets.QMessageBox.information(self, "Nothing to export", "Run a prediction first."); 
            return

        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Export predictions", "", "Excel (*.xlsx);;CSV (*.csv)"
        )
        if not path:
            return

        try:
            if path.lower().endswith(".csv"):
                # CSV = just the long-form predictions
                self.pred_df.to_csv(path, index=False)
            else:
                # Excel = long-form + nicely formatted PredUV_byConc sheet
                from openpyxl.utils import get_column_letter
                with pd.ExcelWriter(path, engine="openpyxl") as w:
                    # Sheet 1 — long-form rows
                    self.pred_df.to_excel(w, sheet_name="LOO_Predictions", index=False)

                    # Sheet 2 — predicted UV by concentration
                    piv = self._predicted_uv_by_concentration(self.pred_df)
                    piv.to_excel(w, sheet_name="PredUV_byConc", index=False)

                    # ---- Light formatting for readability ----
                    wb = w.book
                    if "PredUV_byConc" in wb.sheetnames:
                        ws = wb["PredUV_byConc"]

                        # Freeze top row + first column
                        ws.freeze_panes = "B2"

                        # Reasonable column widths
                        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
                            header_text = col[0].value
                            width = 10
                            if header_text == "Wavelength":
                                width = 12
                            # widen if concentration header is long
                            if isinstance(header_text, str) and len(header_text) > 10:
                                width = min(max(width, len(header_text) + 2), 20)
                            ws.column_dimensions[get_column_letter(col_idx)].width = width

                        # Number format for data cells (6 decimals)
                        for row in ws.iter_rows(min_row=2, min_col=2):
                            for cell in row:
                                cell.number_format = "0.000000"

                    # Optional: light formatting for first sheet too
                    ws1 = w.book["LOO_Predictions"]
                    ws1.freeze_panes = "A2"

            QtWidgets.QMessageBox.information(self, "Exported", f"Saved to:\n{path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export error", str(e))


    def show_spectra_plots(self):
        if self.pred_df is None or self.pred_df.empty:
            QtWidgets.QMessageBox.information(self, "No predictions", "Run 'Predict ALL' first."); return
        df = self.pred_df.copy()
        if "Concentration" not in df.columns:
            QtWidgets.QMessageBox.information(self, "Need concentration", "Pick a Concentration column before predictions."); return

        for col in ["Wavelength","A_ISARS","Concentration"]:
            if col in df.columns: df[col] = pd.to_numeric(df[col], errors="coerce")
        df = df.dropna(subset=["Wavelength","Concentration"])
        actual_piv = (df.pivot_table(index="Wavelength", columns="Concentration", values="UV-Vis_actual", aggfunc="first")
                        .sort_index().sort_index(axis=1))
        pred_piv = (df.pivot_table(index="Wavelength", columns="Concentration", values="UV-Vis_pred", aggfunc="first")
                        .sort_index().sort_index(axis=1))
        comp = df[["UV-Vis_actual","UV-Vis_pred"]].dropna(); have_actual = not comp.empty

        dlg = QtWidgets.QDialog(self); dlg.setWindowTitle("Spectra (per concentration)"); dlg.resize(980, 960)
        lay = QtWidgets.QVBoxLayout(dlg)
        canvas = MplCanvas(self, w=8.0, h=10.0, dpi=120); lay.addWidget(canvas)

        ax1 = canvas.fig.add_subplot(311)
        ax2 = canvas.fig.add_subplot(312, sharex=ax1)
        ax3 = canvas.fig.add_subplot(313)

        ax1.grid(True, alpha=0.6); ax1.set_title("Actual UV-Vis vs Wavelength"); ax1.set_xlabel("Wavelength"); ax1.set_ylabel("UV-Vis (actual)")
        if (actual_piv.notna().sum().sum()) > 0:
            for conc in actual_piv.columns:
                y = actual_piv[conc]
                if y.notna().any(): ax1.plot(actual_piv.index, y, linewidth=2.0, label=f"{conc:g}")
            ax1.legend(title="Concentration", ncols=4, fontsize=8, loc="upper left", bbox_to_anchor=(0,1.02))
        else:
            ax1.text(0.5,0.5,"No actual values", transform=ax1.transAxes, ha="center", va="center", color="#64748B")

        ax2.grid(True, alpha=0.6); ax2.set_title("Predicted UV-Vis vs Wavelength"); ax2.set_xlabel("Wavelength"); ax2.set_ylabel("UV-Vis (pred)")
        if (pred_piv.notna().sum().sum()) > 0:
            for conc in pred_piv.columns:
                y = pred_piv[conc]
                if y.notna().any(): ax2.plot(pred_piv.index, y, linewidth=2.0, label=f"{conc:g}")
            ax2.legend(title="Concentration", ncols=4, fontsize=8, loc="upper left", bbox_to_anchor=(0,1.02))
        else:
            ax2.text(0.5,0.5,"No predictions", transform=ax2.transAxes, ha="center", va="center", color="#64748B")

        ax3.grid(True, alpha=0.6); ax3.set_title("Predicted vs Actual"); ax3.set_xlabel("Actual UV-Vis"); ax3.set_ylabel("Predicted UV-Vis")
        if have_actual:
            x = comp["UV-Vis_actual"].to_numpy(); y = comp["UV-Vis_pred"].to_numpy()
            ax3.scatter(x, y, s=26, alpha=0.85)
            lo = float(np.nanmin([x.min(), y.min()])); hi = float(np.nanmax([x.max(), y.max()]))
            ax3.plot([lo, hi], [lo, hi], linewidth=2.0)
            r2 = centered_r2(x, y); rmse = float(np.sqrt(np.mean((y-x)**2)))
            ax3.text(0.02,0.98, f"R²={r2:.4f}\nRMSE={rmse:.6g}", transform=ax3.transAxes, ha="left", va="top", color="#0F172A")
        else:
            ax3.text(0.5,0.5,"Need actual values to compare", transform=ax3.transAxes, ha="center", va="center", color="#64748B")

        canvas.draw(); dlg.exec()

    # ---------- Save helpers ----------
    def _save_table(self, df: pd.DataFrame, sheet: str):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, f"Export {sheet}", "", "Excel (*.xlsx);;CSV (*.csv)")
        if not path: return
        ext = Path(path).suffix.lower()
        try:
            if ext == ".csv": df.to_csv(path, index=False)
            else:
                with pd.ExcelWriter(path, engine="openpyxl") as w:
                    df.to_excel(w, sheet_name=sheet, index=False)
            QtWidgets.QMessageBox.information(self, "Exported", f"Saved to:\n{path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Export error", str(e))

# ---------------- main ----------------
if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    apply_light_hicontrast_theme(app)
    win = MainWindow()
    win.show()
    app.exec()